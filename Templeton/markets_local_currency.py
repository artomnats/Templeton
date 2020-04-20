#!/usr/bin/env python3.6
"""
Implemented by Artavazd Mnatsakanyan

Description:

    Get all market fund percents from given urls and dump to csv file.

Dependencies & Supported versions:

    Python 3.6.x

Libraries:

    os, sys, re, json, time, argparsem, requests, configparser, enum, bs4, time, selenium

Revision:
    v0.1 alpha (19/04/2020)
        Initial version
Usage:
    ./market_local_currency.py -c config.ini
"""
try:
    import os
    import re
    import sys
    import lxml
    import logging
    import argparse
    import requests
    import openpyxl
    import configparser
    from enum import Enum
    from time import sleep
    from bs4 import BeautifulSoup
    from selenium import webdriver
    from datetime import datetime, timedelta
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.common.exceptions import TimeoutException
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.common.exceptions import NoSuchElementException
    from selenium.webdriver.support import expected_conditions as EC
except ImportError as exception:
    print ('%s - Please install the necessary libraries.' % exception)
    sys.exit(1)

SECTION = Enum('SECTION', (('default')))
OPTIONS = Enum('OPTIONS', (('loglevel'), ('log'), ('urls'), ('excel_file_name')))

Log = Enum('Log', (('INFO'), ('WARNING'), ('ERROR'), ('CRITICAL'), ('DEBUG')), start=0)
# Logger name
logger = logging.getLogger('local_currencies')


def check_internet_connection():
    try:
        requests.get('http://216.58.192.142', timeout=1)
        return True
    except ConnectionError as ce:
        return False
    except Exception:
        return False


def get_configurations(file_name):
    """
    Read file content and create dictionary which with options
    Args:
        file_name - the configuration file
    Returns:
    config - a dict with configuration options
    """
    Config = configparser.ConfigParser()
    Config.read(file_name)
    config = {}
    for each_section in Config.sections():
        config[each_section] = dict((eachKey, eachValue) for eachKey, eachValue in Config.items(each_section))
    return config


def print_log_msg(msg, level):
    """
    Print and add the message into the log file.
    """
    # Check the log level
    {
        Log.INFO.value: lambda msg: logger.info(msg),
        Log.DEBUG.value: lambda msg: logger.debug(msg),
        Log.WARNING.value: lambda msg: logger.warning(msg),
        Log.ERROR.value: lambda msg: logger.error(msg),
        Log.CRITICAL.value: lambda msg: logger.critical(msg)
    }[level](msg)

    print(msg)


def get_local_currencies(urls, excel_file):
    """
    Login to given url, find chat, get chat messages for 6 month, dump all messages to text file.
    Args:
        urls - all urls from config file for getting information
        excel_file - excel file name from config file for dumping scraped info.
    Returns:
        result - all information from web pages.
    """
    pattern_for_date = r'\s+(\d+/\d+/\d+)\s+'
    result = {}
    all_market_currencies = {}
    try:
        # Chrome Options
        chrome_options = webdriver.ChromeOptions()

#        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--lang=en-US')
        chrome_options.add_argument('--dns-prefetch-disable')
        chrome_options.add_experimental_option("prefs", {"download.default_directory": os.getcwd()})

        browser = webdriver.Chrome(chrome_options=chrome_options)
        # makes sure slower connections work as well
        browser.implicitly_wait(10)

        for url in urls:
            result.setdefault(url,{})
            print_log_msg('Open %s url for getting market data' % url, Log.DEBUG.value)
            browser.get(url)

            #sections = browser.find_elements_by_xpath("//div[@class='fund-tabs__tab au-target']")
            print_log_msg('Find Portfolio section and click', Log.DEBUG.value)
            sections = browser.find_element_by_xpath("//ul[@class='nav nav-justified fund-tabs au-target']").find_elements_by_tag_name('li')

            for section in sections:
                if section.text == 'Portfolio':
                    section.click()
                    break
            sleep(3)
            timeout = 2
            try:
                element_present = EC.presence_of_element_located((By.ID, 'main'))
                WebDriverWait(browser, timeout).until(element_present)
            except TimeoutException:
                print("Timed out waiting for page to load")
            finally:
                print("Page loaded")

            all_market_currencies = {}

            sleep(2)
            print_log_msg('Try to get all percents for market table', Log.DEBUG.value)
            if browser.find_elements_by_xpath("//div[@class='fti-chart au-target']"):

                for market_table in browser.find_elements_by_xpath("//div[@class='fti-chart au-target']"):

                    prices = []
                    print_log_msg('Get FUND name', Log.DEBUG.value)
                    currency_info = market_table.find_element_by_xpath("div[@class='row']").text
                    # Get name from currency info
                    name = currency_info.split('\n')[0]
                    # Get date from currency info
                    date = re.search(pattern_for_date, currency_info).group(0)

                    print_log_msg('Try to find percents for each FUND type', Log.DEBUG.value)
                    for price in market_table.find_element_by_tag_name("table").find_elements(By.TAG_NAME, "tr"):
                        if len(price.text) > 1 and price.text != '  FUND':
                            price_name = price.text.split('\n')[0]
                            price_percent = price.text.split('\n')[-1]
                            price = price_name + ' ' + price_percent
                            prices.append(price)
                    all_market_currencies[name + ' ' + date] = ', '.join(prices)

            elif browser.find_elements_by_xpath("//div[@class='au-target col-sm-12']"):

                for market_table in browser.find_elements_by_xpath("//div[@class='au-target col-sm-12']"):

                    prices = []
                    print_log_msg('Get FUND name', Log.DEBUG.value)
                    currency_info = market_table.find_element_by_xpath("div[@class='row']").text
                    # Get name from currency info
                    name = currency_info.split('\n')[0]
                    # Get date from currency info
                    date = re.search(pattern_for_date, currency_info).group(0)

                    print_log_msg('Try to find percents for each FUND type', Log.DEBUG.value)
                    for price in market_table.find_element_by_tag_name("table").find_elements(By.TAG_NAME, "tr"):
                        if len(price.text) > 1 and price.text != '  FUND':
                            price_name = price.text.split('\n')[0]
                            price_percent = price.text.split('\n')[-1]
                            price = price_name + ' ' + price_percent
                            prices.append(price)
                    all_market_currencies[name + ' ' + date] = ', '.join(prices)

            result[url] = all_market_currencies

    except Exception as exception:
        print (exception)

    finally:
#        browser.delete_all_cookies()
        browser.close()

    if result:
        dump_to_csv_file(result, excel_file)


def dump_to_csv_file(result, excel_file):
    """
    Dump all fund peresents to csv file.
    Args:
        result - all information from market tables.
        excel_file - excel file name from config file
    Returns:
        ---
    """
    try:
        file_name = excel_file
        csv_file = openpyxl.load_workbook(file_name)
        ws = csv_file.active
        print_log_msg("File %s opened successfully" % file_name, Log.INFO.value)
    except Exception as exception:
        print_log_msg("%s file does not exist" % file_name, Log.ERROR.value)
        sys.exit(0)
    index = 2
    for url, market_table in result.items():
        for fund_type, fund_percent in market_table.items():
            ws['A' + str(index)] = fund_type
            ws['B' + str(index)] = fund_percent
            ws['C' + str(index)] = url
            index = index + 1

    csv_file.save(file_name)


def main():
    """
    Main function
    """
    """
    Config file passing as script argument
    """
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--config_file', type=str, help='Configuration file')
    args = parser.parse_args()

    if not args.config_file:
        print("Configuration file passed to the script as an argument does not exist.")
        sys.exit(0)

    configs = get_configurations(args.config_file)

    # For logging debug, warning, error and info messages into log file
    log_file = configs.get(SECTION.default.name).get(OPTIONS.log.name)
    level = configs.get(SECTION.default.name).get(OPTIONS.loglevel.name)
    logging.basicConfig(filename=log_file, level=level, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            filemode='w')

    # Check internet connection, than go further
    if not check_internet_connection():
        print_log_msg('Please check internet connection, than try again', Log.ERROR.value)
        sys.exit(0)

    # Get urls from config file
    urls = configs.get(SECTION.default.name).get(OPTIONS.urls.name)
    # Get excel file name from config file
    excel_file = configs.get(SECTION.default.name).get(OPTIONS.excel_file_name.name)

    get_local_currencies(urls.split(', '), excel_file)

if __name__ == "__main__":
    main()
